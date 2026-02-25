from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Any
from models import Nurse, ShiftType

class Exporter:
    @staticmethod
    def export_to_excel(nurses: List[Nurse], year: int, month: int, filename: str):
        wb = Workbook()
        ws = wb.active
        ws.title = f"Schedule {month}-{year}"
        
        # Headers
        headers = ["No.", "Name"] + [str(d) for d in range(1, 32)] + ["M Hours", "E Hours", "N Hours", "Total Hours"]
        ws.append(headers)
        
        # Row Data (Placeholder)
        for i, nurse in enumerate(nurses, 1):
            row = [i, nurse.name]
            # Days 1-31
            for d in range(1, 32):
                shift = nurse.get_shift(d)
                row.append(shift.value if shift != ShiftType.OFF else "")
            
            # Totals (as hours: count × shift_hours)
            m_hrs = nurse.count_shift_type(ShiftType.MORNING) * 6
            e_hrs = nurse.count_shift_type(ShiftType.EVENING) * 6
            n_hrs = nurse.count_shift_type(ShiftType.NIGHT) * 12
            row.append(m_hrs)
            row.append(e_hrs)
            row.append(n_hrs)
            row.append(m_hrs + e_hrs + n_hrs)
            
            ws.append(row)

        # --- Daily Totals ---
        # Calculate totals per column (day)
        total_m = ["Total M", ""]
        total_e = ["Total E", ""]
        total_n = ["Total N", ""]
        
        for d in range(1, 32):
            m_count = 0
            e_count = 0
            n_count = 0
            for nurse in nurses:
                shift = nurse.get_shift(d)
                if shift == ShiftType.MORNING: m_count += 1
                elif shift == ShiftType.EVENING: e_count += 1
                elif shift == ShiftType.NIGHT: n_count += 1
            
            total_m.append(m_count)
            total_e.append(e_count)
            total_n.append(n_count)
            
        ws.append([]) # Empty row
        ws.append(total_m)
        ws.append(total_e)
        ws.append(total_n)

        from openpyxl.styles import Font
        # Style the totals (optional but nice)
        for row_idx in range(ws.max_row - 2, ws.max_row + 1):
            for cell in ws[row_idx]:
                cell.font = Font(bold=True)

        wb.save(filename)
        print(f"Exported to {filename}")
